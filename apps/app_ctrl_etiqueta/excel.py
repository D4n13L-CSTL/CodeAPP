from openpyxl import Workbook
from openpyxl.styles import Font, Color
from openpyxl.styles import Border, Side
from openpyxl.styles import Font, PatternFill, Border, Side
import sqlite3
from pathlib import Path
import io


base_dir  = Path().resolve()
db_path = base_dir / "DATABASE.db"

def get_db():
    conexion = sqlite3.connect(str(db_path))
    return conexion 

class Buscar_fecha ():
    def buscar(self, fecha_init,fecha_end):
        self.conexion = get_db()
        self.cursor = self.conexion.cursor()
        self.cursor.execute(f"""SELECT * FROM Ctrl_Etiquetado WHERE FECHA BETWEEN  ? AND ?""", (fecha_init,fecha_end))
        columns = [column[0] for column in self.cursor.description]
        rows = self.cursor.fetchall()
        data = [dict(zip(columns, row)) for row in rows]
        self.conexion.close()
        return data
    


def generar_excel(resultados):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mi Hoja"

    # Colores
    color_encabezado = "ADD8E6"  # Azul claro
    color_fila_azul = "DDEBF7"   # Azul suave
    color_fila_gris = "F2F2F2"   # Gris claro

    # Borde delgado
    borde_delgado = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Agregar encabezados con estilo
    encabezado = ["CODIGO", "PROVEEDOR", "CONTADO POR", "CANTIDAD ETIQUETAS", "CANTIDAD UNIDADES", "RESPONSABLE", "FECHA", "CONTENEDOR"]
    ws.append(encabezado)  
    for cell in ws[1]:  
        cell.font = Font(bold=True, color="000000")  # Negrita y texto negro
        cell.fill = PatternFill(start_color=color_encabezado, end_color=color_encabezado, fill_type="solid")
        cell.border = borde_delgado  # Aplicar borde

    # Agregar datos con estilos alternos
    for index, i in enumerate(resultados, start=2):  # Empezamos desde la segunda fila
        ws.append(list(i.values())[1:9])
        fill_color = color_fila_azul if index % 2 == 0 else color_fila_gris  # Alternar colores
        
        for cell in ws[index]:
            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            cell.border = borde_delgado  # Aplicar borde a todas las celdas

    ws.auto_filter.ref = ws.dimensions  # Activa el filtro en todo el rango de celdas usadas

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Guardar el archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output